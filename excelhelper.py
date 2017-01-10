import json
import os
import logging
import requests
#import ast

from openpyxl import Workbook

#logging.basicConfig(filename='info.log', level=logging.DEBUG)


def readJson (jsonpath):
    job_type_lists = []
    with open(jsonpath,'r',encoding='utf-8') as f:
        print(f)
        for each_line in f.readlines():
            try:
                json_obj = json.loads(each_line)
                job_type_lists.append(json_obj)
                #return json_obj
            except:
                pass
    return job_type_lists

def write_excel_xueqiu_friends (json_obj_list, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "followers"

    rownum = 3
    colnum = 1

    #get head
    headlist=[]
    if len(json_obj_list)>0:
        for xx in json_obj_list[0].get('followers')[0]:
            headlist.append(xx)
    print(headlist)

    for json_obj in json_obj_list:
        totalcount = json_obj.get('count')    
        ws.cell(row=1, column=1).value = totalcount
        friends = json_obj.get('followers')
        
        #get head

        for friends_info in friends:
            colnum = 1
            #print(friends_info)
            for friends_info_item in headlist:
                #stritem = friends_info_item + ''
                ws.cell(row=2, column=colnum).value = friends_info_item
                ws.cell(row=rownum, column=colnum).value = friends_info[friends_info_item]
                colnum += 1
            rownum+=1
    wb.save('D:/ma/python/xueqiu/' + filename + '.xlsx')
            
if __name__ == '__main__':
    #logging.info('start generating Excel file...')
    #process('D:/LagouJobInfo/lagou')
    #test = requests.get('http://www.baidu.com')
    #print(test)
    xueqiu_json = readJson('D:/ma/python/xueqiu/pofriends.json')
    write_excel_xueqiu_friends(xueqiu_json,'xueqiu_followers')
    #logging.info('Done! Please check your result...')
